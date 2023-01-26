import openpyxl

wb = openpyxl.load_workbook('./pds_source.xlsx') #엑셀 파일 오픈
print('pds_source.xlsx open...')
f = open("./pds.txt", 'w') #메모장 오픈
f.write('{\n') #json 입력 시작

#정보 입력
grade = int(input("학년: ")) #학년 입력
semester = int(input("학기: ")) #학기 입력

sheet = wb.active

parts = wb.sheetnames #sheetname 저장
for i in parts: #각 sheet 탐색
    # print(i)
    partName = wb[i].cell(row=1, column = 1).value #분류 시트 이름
    partName=partName.replace(" ","")
    if(wb[i].cell(row=3, column=1).value == None): #단원 입력 셀이 아니면 가로로 탐색
        for j in wb[i].iter_rows(min_row=1):
            print('')
                
    else: #단원 입력 셀 경우 세로로 탐색
        for j in wb[i].iter_cols(min_col=4):
            nTail = 1
            for cell in j:
                nBody = cell.column-3
                if(cell.row == 1):
                    title = cell.value #title 저장
                
                if(cell.row == 2):
                    contentType = cell.value #content-type 저장                
                    if(contentType == 'jpg' or contentType=='zip'): #type 저장
                        type = 'blue'
                    else:
                        type = 'red'

                if(cell.value == '○'): #json 입력
                    id = i+'0'+str(nBody)+'000'+str(nTail) #id 저장
                    nTail += 1    
                    unit = wb[i].cell(row=cell.row, column = 2).value #중단원 저장
                    unitName = wb[i].cell(row=cell.row, column = 3).value #중단원명 저장
                    #filename 형식에 맞춰 저장
                    file = title.replace(" ","") 
                    file = file.replace("PDF","")
                    if(contentType == 'zip'):
                        fileName = str(grade)+'_'+str(semester)+'_'+str(lesson)+'_'+str(unit)+'_'+file+'.'+contentType
                    else:
                        fileName = str(lesson)+'_'+str(unit)+'_'+file+'.'+contentType
                    #filename 끝
                    #url 형식에 맞춰 저장
                    tit = title.replace(" ","")
                    if(contentType == 'zip'):
                        url = partName +'/'+tit+'/'+fileName
                    else:
                        url = partName +'/'+tit+'/'+str(lesson)+'단원/'+fileName
                    #url 끝
                    toc1 = str(lesson)+'. '+lessonName #toc-1 입력
                    toc2 = str(unit)+'. '+unitName #toc-2 입력
                    
                    # 코드 입력
                    f.write("\t\""+id+"\" : {\n")
                    f.write("\t\t\"id\" : \""+id+"\",\n")
                    f.write("\t\t\"title\" : \""+title+"\",\n")
                    f.write("\t\t\"url\" : \""+url+"\",\n")
                    f.write("\t\t\"filename\" : \""+fileName+"\",\n")
                    f.write("\t\t\"content-type\" : \""+contentType+"\",\n")
                    f.write("\t\t\"type\" : \""+type+"\",\n")
                    f.write("\t\t\"type-name\" : \""+toc2+"\",\n")
                    f.write("\t\t\"toc-1\" : \""+toc1+"\",\n")
                    f.write("\t\t\"toc-2\" : \""+toc2+"\",\n")
                    f.write("\t\t\"toc-3\" : \"\"\n")
                    f.write("\t\t},\n")

                if(cell.value == '-'):
                    continue

                if(cell.value == None):
                    lesson = wb[i].cell(row=cell.row, column=1).value #대단원 저장
                    lessonName = wb[i].cell(row=cell.row, column=3).value #대단원명 저장
                   
                # print(cell.value)
                
                


f.write('}') #json 입력 끝
f.close()
